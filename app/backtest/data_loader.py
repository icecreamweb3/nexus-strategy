"""K线 CSV/Excel 加载 / OHLC kline data loader."""
import csv
from dataclasses import dataclass
from typing import List

# 常见列名映射（小写）
_COL_ALIASES = {
    "open_time": ["open_time", "time", "date", "datetime", "timestamp", "开盘时间", "时间"],
    "open": ["open", "o", "开盘价"],
    "high": ["high", "h", "最高价"],
    "low": ["low", "l", "最低价"],
    "close": ["close", "c", "收盘价"],
    "volume": ["volume", "vol", "v", "成交量"],
}


@dataclass
class Kline:
    index: int          # 1-based K线索引
    open_time: str
    open: float
    high: float
    low: float
    close: float
    volume: float


class DataLoadError(Exception):
    pass


def _detect_columns(header: List[str]) -> dict:
    lowered = [h.strip().lower() for h in header]
    mapping = {}
    for field, aliases in _COL_ALIASES.items():
        for alias in aliases:
            if alias in lowered:
                mapping[field] = lowered.index(alias)
                break
    return mapping


def _rows_from_csv(path: str):
    with open(path, newline="", encoding="utf-8-sig") as f:
        yield from csv.reader(f)


def _rows_from_xlsx(path: str):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield ["" if c is None else c for c in row]


def load_klines(path: str) -> List[Kline]:
    """加载 K线文件，自动识别表头与列名。无表头时按 time,open,high,low,close,volume 顺序解析。"""
    rows = _rows_from_xlsx(path) if path.lower().endswith(".xlsx") else _rows_from_csv(path)

    klines: List[Kline] = []
    mapping = None
    for raw in rows:
        if not raw or all(str(c).strip() == "" for c in raw):
            continue
        if mapping is None:
            candidate = _detect_columns([str(c) for c in raw])
            if "open" in candidate and "close" in candidate:
                mapping = candidate
                continue
            # 无表头：按固定顺序
            mapping = {name: i for i, name in enumerate(
                ["open_time", "open", "high", "low", "close", "volume"]) if i < len(raw)}
        try:
            klines.append(Kline(
                index=len(klines) + 1,
                open_time=str(raw[mapping["open_time"]]) if "open_time" in mapping else "",
                open=float(raw[mapping["open"]]),
                high=float(raw[mapping["high"]]),
                low=float(raw[mapping["low"]]),
                close=float(raw[mapping["close"]]),
                volume=float(raw[mapping["volume"]]) if "volume" in mapping else 0.0,
            ))
        except (ValueError, IndexError, KeyError):
            continue  # 跳过无法解析的行

    if not klines:
        raise DataLoadError("no valid kline rows parsed")
    return klines
